from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook('mock_data.xlsx')
ws = wb.active
# this is where we haev loaded the data 
# Calculate total salary (assuming salary is in column E, skipping header)
total_salary = 0
for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
    for cell in row:
        total_salary += cell.value

print(f"Total Salary: {total_salary}")
